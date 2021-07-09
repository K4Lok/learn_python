from openpyxl import load_workbook

wb = load_workbook('yourExcel.xlsx')

ws2 = wb['Dataset1'] #Get worksheet by it's name

file = open('objectData.txt', 'w+')

for row in ws2.iter_rows(min_col=1, min_row=2, max_col=2, max_row=ws2.max_row-1):
    file.write('{week: "' + str(row[0].value) + '", count: '+ str(row[1].value) + '},\n')

file.close()
