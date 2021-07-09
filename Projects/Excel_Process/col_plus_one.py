from openpyxl import Workbook

wb = load_workbook('yourExcelFile.xlsx')

ws2 = wb.get_sheet_by_name("Dataset1")

start_row = 2
end_row = 29
increase_number = 1

for row_index in range(start_row,end_row-1):
    ws2.cell(row=row_index,column=1).value = str(int(ws2.cell(row=row_index, column=1).value) + increase_number)

wb.save('yourExcelFile.xlsx')
