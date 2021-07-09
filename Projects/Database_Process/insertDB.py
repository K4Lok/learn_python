import mysql.connector
from openpyxl import load_workbook

mydb = mysql.connector.connect(
    host="localhost/IP",
    user="userName",
    password="password",
    database="yourDB"
)

print("Connected!")

cur = mydb.cursor()

sql = "INSERT INTO yourTable (Nmae, address) VALUES (%s, %s)"

wb = load_workbook("yourExcelFile.xlsx")

ws2 = wb["Dataset1"] #Name of the worksheet

for rows in ws2.iter_rows(min_col=1, min_row=2, max_col=2, max_row=ws2.max_row-1):
    #my_str = "({}, {})".format(rows[0].value, rows[1].value)
    #print(my_str)
    cur.execute(sql, (rows[0].value, rows[1].value))

mydb.commit() #Won't insert any datas without committing

print("Finished!")
